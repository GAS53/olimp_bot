import os

from openpyxl import load_workbook

import settings


def get_files():
    li = []
    for root, _, files in os.walk(settings.TEXT_FILES):
        for file in files:
            f_path = os.path.join(root, file)
            li.append(f_path)
            print(f_path)
    return li


def prepate():
    li = get_files()
    res = {}
    for file in li:
        res.update(pars_excel_file(file))
    return res


def pars_row(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]


def cleaner(val: str):
    if val:
        return val.strip().replace('\n', '').lower()
    return None
    


def pars_excel_file(file):
    print(f' паршу файл {file}')
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    di = {}
    for row in pars_row(ws):
        di[cleaner(row[0])] = cleaner(row[1])
    return di


def finder(di, find):
    res = ''
    find = cleaner(find)
    
    for i in di.keys():
        if i and find in i:
            res += f'{di[i]}\n'
    return res
        


if __name__ == '__main__':
    prepate()